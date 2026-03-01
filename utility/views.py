from django.http import JsonResponse
from django.shortcuts import render, redirect
from .models import Site
from .forms import InspectionForm, DieselFillingForm
from .forms import SiteUploadForm

def site_search_api(request):
    q = request.GET.get("q", "").strip()
    qs = Site.objects.all().order_by("site_name")
    if q:
        qs = qs.filter(site_name__icontains=q)
    qs = qs[:20]

    results = [{
        "id": s.id,
        "text": f"{s.site_name} ({s.site_code or ''})",
        "site_code": s.site_code,
        "bss_incharge_name": s.bss_incharge_name,
        "bss_incharge_mobile": s.bss_incharge_mobile,
        "technician_name": s.technician_name,
        "technician_mobile": s.technician_mobile,
    } for s in qs]

    return JsonResponse({"results": results})

def inspection_create(request):
    if request.method == "POST":
        form = InspectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Inspection saved successfully.")
            return redirect("inspection_create")
    else:
        form = InspectionForm()
    return render(request, "utility/inspection_form.html", {"form": form})

def diesel_filling_create(request):
    if request.method == "POST":
        form = DieselFillingForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Diesel filling record saved successfully.")
            return redirect("diesel_filling_create")
    else:
        form = DieselFillingForm()
    return render(request, "utility/diesel_filling_form.html", {"form": form})

import calendar
import os
import tempfile

from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Max, Sum
from django.utils import timezone

import openpyxl

from .models import Site, Inspection, DieselFilling
from .forms import InspectionForm, DieselFillingForm
from .forms_upload import SiteUploadForm
from .diesel_statement import generate_diesel_statement, summarize_diesel_statement


def site_search_api(request):
    q = request.GET.get("q", "").strip()
    qs = Site.objects.all().order_by("site_name")
    if q:
        qs = qs.filter(site_name__icontains=q)
    qs = qs[:20]

    results = [{
        "id": s.id,
        "text": f"{s.site_name} ({s.site_code or ''})",
        "site_code": s.site_code,
        "bss_incharge_name": s.bss_incharge_name,
        "bss_incharge_mobile": s.bss_incharge_mobile,
        "technician_name": s.technician_name,
        "technician_mobile": s.technician_mobile,
    } for s in qs]

    return JsonResponse({"results": results})


def inspection_create(request):
    if request.method == "POST":
        form = InspectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Inspection saved successfully.")
            return redirect("inspection_create")
    else:
        form = InspectionForm()

    return render(request, "utility/inspection_form.html", {"form": form})


def diesel_filling_create(request):
    if request.method == "POST":
        form = DieselFillingForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Diesel filling record saved successfully.")
            return redirect("diesel_filling_create")
    else:
        form = DieselFillingForm()

    return render(request, "utility/diesel_filling_form.html", {"form": form})


def _parse_month_value(month_value):
    if month_value is None:
        return None

    value = str(month_value).strip()
    if not value:
        return None

    if value.isdigit():
        month_num = int(value)
        return month_num if 1 <= month_num <= 12 else None

    lookup = {}
    for idx in range(1, 13):
        lookup[calendar.month_name[idx].lower()] = idx
        lookup[calendar.month_abbr[idx].lower()] = idx

    return lookup.get(value.lower())


def _latest_closed_month(today):
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


def inspection_list(request):
    today = timezone.localdate()
    selected_year = today.year
    selected_month = today.month

    year_raw = (request.GET.get("year") or "").strip()
    month_raw = (request.GET.get("month") or "").strip()

    if year_raw:
        try:
            selected_year = int(year_raw)
        except ValueError:
            messages.error(request, "Invalid year selected. Showing current year.")
            selected_year = today.year

    if month_raw:
        parsed_month = _parse_month_value(month_raw)
        if parsed_month:
            selected_month = parsed_month
        else:
            messages.error(request, "Invalid month selected. Showing current month.")
            selected_month = today.month

    inspections_qs = (
        Inspection.objects.select_related("site")
        .filter(inspection_date__year=selected_year, inspection_date__month=selected_month)
        .order_by("-inspection_date", "site__site_name")
    )

    site_summary = (
        inspections_qs.values(
            "site__site_name",
            "site__site_code",
            "site__bss_incharge_name",
            "site__technician_name",
        )
        .annotate(total_inspections=Count("id"), last_inspection_date=Max("inspection_date"))
        .order_by("site__site_name")
    )

    year_choices = [d.year for d in Inspection.objects.dates("inspection_date", "year", order="DESC")]
    if selected_year not in year_choices:
        year_choices.append(selected_year)
        year_choices = sorted(set(year_choices), reverse=True)

    month_choices = [
        {"number": i, "name": calendar.month_name[i]}
        for i in range(1, 13)
    ]

    context = {
        "inspections": inspections_qs,
        "site_summary": site_summary,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "year_choices": year_choices or [today.year],
        "month_choices": month_choices,
        "total_records": inspections_qs.count(),
        "total_sites": site_summary.count(),
    }
    return render(request, "utility/inspection_list.html", context)


def diesel_records_list(request):
    today = timezone.localdate()
    selected_year = today.year
    selected_month = today.month

    year_raw = (request.GET.get("year") or "").strip()
    month_raw = (request.GET.get("month") or "").strip()

    if year_raw:
        try:
            selected_year = int(year_raw)
        except ValueError:
            messages.error(request, "Invalid year selected. Showing current year.")
            selected_year = today.year

    if month_raw:
        parsed_month = _parse_month_value(month_raw)
        if parsed_month:
            selected_month = parsed_month
        else:
            messages.error(request, "Invalid month selected. Showing current month.")
            selected_month = today.month

    diesel_qs = (
        DieselFilling.objects.select_related("site")
        .filter(date_of_filling__year=selected_year, date_of_filling__month=selected_month)
        .order_by("-date_of_filling", "site__site_name")
    )

    site_summary = (
        diesel_qs.values(
            "site__site_name",
            "site__site_code",
            "site__bss_incharge_name",
            "site__technician_name",
        )
        .annotate(
            total_entries=Count("id"),
            total_diesel_filled=Sum("diesel_filled"),
            last_filling_date=Max("date_of_filling"),
        )
        .order_by("site__site_name")
    )
    total_diesel_filled = diesel_qs.aggregate(total=Sum("diesel_filled"))["total"] or 0

    year_choices = [d.year for d in DieselFilling.objects.dates("date_of_filling", "year", order="DESC")]
    if selected_year not in year_choices:
        year_choices.append(selected_year)
        year_choices = sorted(set(year_choices), reverse=True)

    month_choices = [
        {"number": i, "name": calendar.month_name[i]}
        for i in range(1, 13)
    ]

    context = {
        "records": diesel_qs,
        "site_summary": site_summary,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "year_choices": year_choices or [today.year],
        "month_choices": month_choices,
        "total_records": diesel_qs.count(),
        "total_sites": site_summary.count(),
        "total_diesel_filled": total_diesel_filled,
    }
    return render(request, "utility/diesel_records_list.html", context)


def diesel_statement_list(request):
    today = timezone.localdate()
    latest_year, latest_month = _latest_closed_month(today)
    selected_year = latest_year
    selected_month = latest_month
    selected_bss_incharge = (request.GET.get("bss_incharge") or "").strip()
    selected_site_id = (request.GET.get("site") or "").strip()

    year_raw = (request.GET.get("year") or "").strip()
    month_raw = (request.GET.get("month") or "").strip()

    if year_raw:
        try:
            selected_year = int(year_raw)
        except ValueError:
            messages.error(request, "Invalid year selected. Showing latest allowed month.")
            selected_year = latest_year

    if month_raw:
        parsed_month = _parse_month_value(month_raw)
        if parsed_month:
            selected_month = parsed_month
        else:
            messages.error(request, "Invalid month selected. Showing latest allowed month.")
            selected_month = latest_month

    if (selected_year, selected_month) >= (today.year, today.month):
        selected_year, selected_month = latest_year, latest_month
        messages.error(
            request,
            "Diesel statement can be prepared only for completed months. Current and future months are restricted.",
        )

    rows = generate_diesel_statement(
        month=selected_month,
        year=selected_year,
        bss_incharge=selected_bss_incharge or None,
    )

    if selected_site_id.isdigit():
        rows = [row for row in rows if row["site_id"] == int(selected_site_id)]

    summary_rows = summarize_diesel_statement(rows)

    inspection_years = [d.year for d in Inspection.objects.dates("inspection_date", "year", order="DESC")]
    filling_years = [d.year for d in DieselFilling.objects.dates("date_of_filling", "year", order="DESC")]
    year_choices = sorted(
        {year for year in (inspection_years + filling_years + [latest_year]) if year <= latest_year},
        reverse=True,
    )

    month_choices = [{"number": i, "name": calendar.month_name[i]} for i in range(1, 13)]
    bss_incharge_choices = (
        Site.objects.exclude(bss_incharge_name__isnull=True)
        .exclude(bss_incharge_name__exact="")
        .values_list("bss_incharge_name", flat=True)
        .distinct()
        .order_by("bss_incharge_name")
    )
    site_choices_qs = Site.objects.all()
    if selected_bss_incharge:
        site_choices_qs = site_choices_qs.filter(bss_incharge_name__iexact=selected_bss_incharge)
    site_choices = site_choices_qs.order_by("site_name")

    context = {
        "rows": rows,
        "summary_rows": summary_rows,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "selected_bss_incharge": selected_bss_incharge,
        "selected_site_id": selected_site_id,
        "year_choices": year_choices or [today.year],
        "month_choices": month_choices,
        "current_year": today.year,
        "current_month": today.month,
        "bss_incharge_choices": bss_incharge_choices,
        "site_choices": site_choices,
        "total_sites": len(rows),
        "total_incharges": len(summary_rows),
    }
    return render(request, "utility/diesel_statement_list.html", context)


#@login_required
def upload_sites(request):
    """
    Upload an Excel (.xlsx) file and import/update Site master table.
    """
    if request.method == "POST":
        form = SiteUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data["file"]
            sheet_name = form.cleaned_data.get("sheet") or None
            header_row = form.cleaned_data.get("header_row") or 1

            ext = os.path.splitext(uploaded_file.name)[1].lower()
            if ext != ".xlsx":
                messages.error(request, "Please upload an .xlsx Excel file only.")
                return redirect("upload_sites")

            # Save file to temp path
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                for chunk in uploaded_file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            try:
                created, updated, skipped = import_sites_from_xlsx(
                    tmp_path, sheet_name=sheet_name, header_row=header_row
                )
                messages.success(
                    request,
                    f"Upload completed ✅ Created: {created}, Updated: {updated}, Skipped: {skipped}"
                )
            except Exception as e:
                messages.error(request, f"Import failed: {e}")
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

            return redirect("upload_sites")
    else:
        form = SiteUploadForm()

    return render(request, "utility/upload_sites.html", {"form": form})


def import_sites_from_xlsx(xlsx_path, sheet_name=None, header_row=1):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # Read header row
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    header_map = {h.lower(): idx for idx, h in enumerate(headers)}

    def get(row_values, *possible_headers):
        for h in possible_headers:
            idx = header_map.get(h.lower())
            if idx is not None and idx < len(row_values):
                v = row_values[idx]
                if v is None:
                    continue
                return str(v).strip() if isinstance(v, str) else v
        return None

    created = updated = skipped = 0

    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

        site_name = get(row, "site_name", "site name", "site")
        if not site_name:
            skipped += 1
            continue

        site_code = get(row, "site_code", "site code", "pkey", "site id")
        bss_incharge_name = get(row, "bss_incharge_name", "bss incharge name", "bss incharge")
        bss_incharge_mobile = get(row, "bss_incharge_mobile", "bss incharge mobile", "bss mobile", "incharge mobile")
        technician_name = get(row, "technician_name", "technician name", "technicial name", "technical name")
        technician_mobile = get(row, "technician_mobile", "technician mobile", "technical mobile", "mobile number")

        obj, is_created = Site.objects.get_or_create(site_name=str(site_name).strip())

        if site_code is not None:
            obj.site_code = str(site_code).strip()
        if bss_incharge_name is not None:
            obj.bss_incharge_name = str(bss_incharge_name).strip()
        if bss_incharge_mobile is not None:
            obj.bss_incharge_mobile = str(bss_incharge_mobile).strip()
        if technician_name is not None:
            obj.technician_name = str(technician_name).strip()
        if technician_mobile is not None:
            obj.technician_mobile = str(technician_mobile).strip()

        obj.save()

        if is_created:
            created += 1
        else:
            updated += 1

    return created, updated, skipped
