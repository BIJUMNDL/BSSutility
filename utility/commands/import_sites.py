from django.core.management.base import BaseCommand, CommandError
from utility.models import Site

try:
    import openpyxl
except ImportError:
    openpyxl = None


class Command(BaseCommand):
    help = "Import Site master data from an Excel .xlsx file into the Site table."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to Excel file (.xlsx)")
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Sheet name (optional). If not given, first sheet is used.",
        )
        parser.add_argument(
            "--header-row",
            type=int,
            default=1,
            help="Header row number (1-based). Default=1",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl is not installed. Run: pip install openpyxl")

        xlsx_path = options["xlsx_path"]
        sheet_name = options["sheet"]
        header_row = options["header_row"]

        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            raise CommandError(f"Could not open Excel file: {xlsx_path}\n{e}")

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise CommandError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        # Read header
        headers = []
        for cell in ws[header_row]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")

        # Build header -> column index mapping
        header_map = {h.lower(): idx for idx, h in enumerate(headers)}

        # Helper to fetch cell value by header key options
        def get(row_values, *possible_headers):
            for h in possible_headers:
                idx = header_map.get(h.lower())
                if idx is not None and idx < len(row_values):
                    v = row_values[idx]
                    if v is None:
                        continue
                    v = str(v).strip() if isinstance(v, str) else v
                    return v
            return None

        created = 0
        updated = 0
        skipped = 0

        # Start reading data rows after header
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

            site_name = get(row, "site_name", "site name", "site")
            if not site_name:
                skipped += 1
                continue

            site_code = get(row, "site_code", "site code", "pkey", "site id")
            bss_incharge_name = get(row, "bss_incharge_name", "bss incharge name", "bss incharge")
            bss_incharge_mobile = get(row, "bss_incharge_mobile", "bss incharge mobile", "bss mobile", "incharge mobile")
            technician_name = get(row, "technician_name", "technician name", "technical name", "technicial name")
            technician_mobile = get(row, "technician_mobile", "technician mobile", "technical mobile", "mobile number")

            # Upsert by site_name
            obj, is_created = Site.objects.get_or_create(site_name=str(site_name).strip())

            obj.site_code = str(site_code).strip() if site_code is not None else obj.site_code
            obj.bss_incharge_name = str(bss_incharge_name).strip() if bss_incharge_name is not None else obj.bss_incharge_name
            obj.bss_incharge_mobile = str(bss_incharge_mobile).strip() if bss_incharge_mobile is not None else obj.bss_incharge_mobile
            obj.technician_name = str(technician_name).strip() if technician_name is not None else obj.technician_name
            obj.technician_mobile = str(technician_mobile).strip() if technician_mobile is not None else obj.technician_mobile

            obj.save()

            if is_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"Import completed ✅  Created: {created}, Updated: {updated}, Skipped(empty site_name): {skipped}"
        ))