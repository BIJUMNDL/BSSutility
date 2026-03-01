import datetime
import calendar
from decimal import Decimal

from django.http import HttpResponse, HttpResponseBadRequest
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from openpyxl import Workbook

from .models import Site, Inspection, DieselFilling
from .diesel_statement import generate_diesel_statement, summarize_diesel_statement


def _naive(value):
    """Strip timezone info so openpyxl can handle datetimes."""
    if isinstance(value, datetime.datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _workbook_response(wb: Workbook, filename: str) -> HttpResponse:
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def _autofit_columns(ws):
    for column in ws.columns:
        col_letter = column[0].column_letter
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 50)


def _format_two_decimals(ws, column_indexes, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for idx in column_indexes:
            cell = row[idx - 1]
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = "0.00"


def _parse_month_value(month_value):
    if month_value is None:
        return None

    value = str(month_value).strip()
    if not value:
        return None

    if value.isdigit():
        month_num = int(value)
        return month_num if 1 <= month_num <= 12 else None

    lookup = {}
    for idx in range(1, 13):
        lookup[calendar.month_name[idx].lower()] = idx
        lookup[calendar.month_abbr[idx].lower()] = idx

    return lookup.get(value.lower())


def _latest_closed_month(today):
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


@login_required
def export_sites_xlsx(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sites"

    ws.append([
        "site_name", "site_code",
        "bss_incharge_name", "bss_incharge_mobile",
        "technician_name", "technician_mobile",
    ])

    for s in Site.objects.all().order_by("site_name"):
        ws.append([
            s.site_name, s.site_code,
            s.bss_incharge_name, s.bss_incharge_mobile,
            s.technician_name, s.technician_mobile,
        ])

    return _workbook_response(wb, "sites.xlsx")


def export_inspections_xlsx(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspections"

    ws.append([
        "inspection_date", "site_name", "site_code",
        "eb_reading", "power_factor", "contract_demand", "dg_kwh_reading", "hour_meter_reading", "diesel_balance",
        "dg_status", "dg_remarks",
        "dc_load_kw",
        "powerplant_max_modules",
        "modules_total", "modules_working", "modules_faulty", "modules_spare",
        "battery_backup_hours", "discharge_test_conducted",
        "cleanliness_of_battery",
        "aviation_lamp_condition", "fire_extinguisher_condition",
        "aircondition_status", "free_cooling_status",
        "overall_site_cleanliness", "spare_unused_items_available",
        "created_at",
    ])

    qs = Inspection.objects.select_related("site").all()

    year_raw = (request.GET.get("year") or "").strip()
    month_raw = (request.GET.get("month") or "").strip()

    if year_raw:
        try:
            qs = qs.filter(inspection_date__year=int(year_raw))
        except ValueError:
            pass

    parsed_month = _parse_month_value(month_raw)
    if parsed_month:
        qs = qs.filter(inspection_date__month=parsed_month)

    qs = qs.order_by("-inspection_date", "-created_at")
    for x in qs:
        ws.append([
            _naive(x.inspection_date),
            x.site.site_name, x.site.site_code,
            x.eb_reading, x.power_factor, x.contract_demand, x.dg_kwh_reading, x.hour_meter_reading, x.diesel_balance,
            x.dg_status, x.dg_remarks,
            x.dc_load_kw,
            x.powerplant_max_modules,
            x.modules_total, x.modules_working, x.modules_faulty, x.modules_spare,
            x.battery_backup_hours, x.discharge_test_conducted,
            x.cleanliness_of_battery,
            x.aviation_lamp_condition, x.fire_extinguisher_condition,
            x.aircondition_status, x.free_cooling_status,
            x.overall_site_cleanliness, x.spare_unused_items_available,
            _naive(x.created_at),
        ])

    filename = "inspections.xlsx"
    if year_raw and parsed_month:
        filename = f"inspections_{year_raw}_{parsed_month:02d}.xlsx"
    elif year_raw:
        filename = f"inspections_{year_raw}.xlsx"
    elif parsed_month:
        filename = f"inspections_{parsed_month:02d}.xlsx"

    return _workbook_response(wb, filename)


def export_diesel_xlsx(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "DieselFilling"

    ws.append([
        "date_of_filling", "site_name", "site_code",
        "balance_on_tank", "diesel_filled",
        "hour_meter_reading", "kwh",
        "balance_after_filling",
        "created_at",
    ])

    qs = DieselFilling.objects.select_related("site").all()

    year_raw = (request.GET.get("year") or "").strip()
    month_raw = (request.GET.get("month") or "").strip()

    if year_raw:
        try:
            qs = qs.filter(date_of_filling__year=int(year_raw))
        except ValueError:
            pass

    parsed_month = _parse_month_value(month_raw)
    if parsed_month:
        qs = qs.filter(date_of_filling__month=parsed_month)

    qs = qs.order_by("-date_of_filling", "-created_at")
    for d in qs:
        ws.append([
            _naive(d.date_of_filling),
            d.site.site_name, d.site.site_code,
            d.balance_on_tank, d.diesel_filled,
            d.hour_meter_reading, d.kwh,
            d.balance_after_filling,
            _naive(d.created_at),
        ])

    filename = "diesel_filling.xlsx"
    if year_raw and parsed_month:
        filename = f"diesel_filling_{year_raw}_{parsed_month:02d}.xlsx"
    elif year_raw:
        filename = f"diesel_filling_{year_raw}.xlsx"
    elif parsed_month:
        filename = f"diesel_filling_{parsed_month:02d}.xlsx"

    return _workbook_response(wb, filename)


def export_diesel_statement_xlsx(request):
    today = timezone.localdate()
    latest_year, latest_month = _latest_closed_month(today)
    selected_year = latest_year
    selected_month = latest_month

    year_raw = (request.GET.get("year") or "").strip()
    month_raw = (request.GET.get("month") or "").strip()
    selected_bss_incharge = (request.GET.get("bss_incharge") or "").strip()
    selected_site_id = (request.GET.get("site") or "").strip()

    if year_raw:
        try:
            selected_year = int(year_raw)
        except ValueError:
            selected_year = latest_year

    parsed_month = _parse_month_value(month_raw)
    if parsed_month:
        selected_month = parsed_month

    if (selected_year, selected_month) >= (today.year, today.month):
        return HttpResponseBadRequest(
            "Diesel statement export is allowed only for completed months. Current and future months are restricted."
        )

    rows = generate_diesel_statement(
        month=selected_month,
        year=selected_year,
        bss_incharge=selected_bss_incharge or None,
    )
    if selected_site_id.isdigit():
        rows = [row for row in rows if row["site_id"] == int(selected_site_id)]
    summary_rows = summarize_diesel_statement(rows)

    wb = Workbook()

    ws = wb.active
    ws.title = "Diesel Statement"
    ws.append([
        "month", "bss_incharge_name", "site_name", "site_code",
        "opening_balance", "opening_hmr", "diesel_filled",
        "closing_balance", "closing_hmr",
        "diesel_consumption", "running_hours", "hourly_consumption",
        "remarks",
    ])
    for row in rows:
        ws.append([
            row["month"],
            row["bss_incharge_name"],
            row["site_name"],
            row["site_code"],
            row["opening_balance"],
            row["opening_hmr"],
            row["diesel_filled"],
            row["closing_balance"],
            row["closing_hmr"],
            row["diesel_consumption"],
            row["running_hours"],
            row["hourly_consumption"],
            row["remarks"],
        ])
    _format_two_decimals(ws, [5, 6, 7, 8, 9, 10, 11, 12])
    _autofit_columns(ws)

    ws2 = wb.create_sheet("BSSIncharge Summary")
    ws2.append([
        "bss_incharge_name", "site_count",
        "diesel_filled_total", "diesel_consumption_total",
        "running_hours_total", "hourly_consumption_total",
    ])
    for row in summary_rows:
        ws2.append([
            row["bss_incharge_name"],
            row["site_count"],
            row["diesel_filled_total"],
            row["diesel_consumption_total"],
            row["running_hours_total"],
            row["hourly_consumption_total"],
        ])
    _format_two_decimals(ws2, [3, 4, 5, 6])
    _autofit_columns(ws2)

    filename = f"diesel_statement_{selected_year}_{selected_month:02d}.xlsx"
    return _workbook_response(wb, filename)


@login_required
def export_all_xlsx(request):
    wb = Workbook()

    # Sheet 1: Sites
    ws1 = wb.active
    ws1.title = "Sites"
    ws1.append([
        "site_name", "site_code",
        "bss_incharge_name", "bss_incharge_mobile",
        "technician_name", "technician_mobile",
    ])
    for s in Site.objects.all().order_by("site_name"):
        ws1.append([s.site_name, s.site_code, s.bss_incharge_name, s.bss_incharge_mobile, s.technician_name, s.technician_mobile])

    # Sheet 2: Inspections
    ws2 = wb.create_sheet("Inspections")
    ws2.append([
        "inspection_date", "site_name", "site_code",
        "eb_reading", "power_factor", "contract_demand", "dg_kwh_reading", "hour_meter_reading", "diesel_balance",
        "dg_status", "dg_remarks",
        "dc_load_kw",
        "powerplant_max_modules",
        "modules_total", "modules_working", "modules_faulty", "modules_spare",
        "battery_backup_hours", "discharge_test_conducted",
        "cleanliness_of_battery",
        "aviation_lamp_condition", "fire_extinguisher_condition",
        "aircondition_status", "free_cooling_status",
        "overall_site_cleanliness", "spare_unused_items_available",
        "created_at",
    ])
    for x in Inspection.objects.select_related("site").all().order_by("-inspection_date", "-created_at"):
        ws2.append([
            _naive(x.inspection_date),
            x.site.site_name, x.site.site_code,
            x.eb_reading, x.power_factor, x.contract_demand, x.dg_kwh_reading, x.hour_meter_reading, x.diesel_balance,
            x.dg_status, x.dg_remarks,
            x.dc_load_kw,
            x.powerplant_max_modules,
            x.modules_total, x.modules_working, x.modules_faulty, x.modules_spare,
            x.battery_backup_hours, x.discharge_test_conducted,
            x.cleanliness_of_battery,
            x.aviation_lamp_condition, x.fire_extinguisher_condition,
            x.aircondition_status, x.free_cooling_status,
            x.overall_site_cleanliness, x.spare_unused_items_available,
            _naive(x.created_at),
        ])

    # Sheet 3: Diesel
    ws3 = wb.create_sheet("DieselFilling")
    ws3.append([
        "date_of_filling", "site_name", "site_code",
        "balance_on_tank", "diesel_filled",
        "hour_meter_reading", "kwh",
        "balance_after_filling",
        "created_at",
    ])
    for d in DieselFilling.objects.select_related("site").all().order_by("-date_of_filling", "-created_at"):
        ws3.append([
            _naive(d.date_of_filling),
            d.site.site_name, d.site.site_code,
            d.balance_on_tank, d.diesel_filled,
            d.hour_meter_reading, d.kwh,
            d.balance_after_filling,
            _naive(d.created_at),
        ])

    return _workbook_response(wb, "bssutility_all.xlsx")
